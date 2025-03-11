import openpyxl
from openpyxl import load_workbook
import customtkinter as ctk
from tkinter import messagebox
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# Initialize the main application window
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

def generate_pdf(data, total_earnings, total_deductions, net_pay):
    try:
        filename = f"PaySlip_{data['Employee Name'].replace(' ', '_')}_{data['Employee ID']}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        elements.append(Paragraph("<u>Monthly Payroll Slip</u>", title_style))

        # Employee Details Table
        details_data = [
            ["Employee Name:", data["Employee Name"]],
            ["Employee ID:", data["Employee ID"]],
            ["Department:", data["Department"]],
            ["Designation:", data["Designation"]],
            ["Date of Joining:", str(data["Date of Joining"])],
            ["Date of Birth:", str(data["Date of Birth"])],
            ["UAN:", data["UAN"]],
            ["PF No:", data["PF No"]],
            ["ESI No:", data["ESI No"]],
        ]
        details_table = Table(details_data, colWidths=[1.5*inch, 3*inch])
        details_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(details_table)
        elements.append(Spacer(1, 0.3*inch))

        # Earnings Table
        earnings_data = [
            ["Basic Salary", f"₹{data['Basic Salary']}"],
            ["Conveyance", f"₹{data['Conveyance']}"],
            ["Special Allowance", f"₹{data['Special Allowance']}"],
            ["", ""],
            ["Total Earnings", f"₹{total_earnings}"]
        ]
        earnings_table = Table(earnings_data, colWidths=[3.5*inch, 1.5*inch])
        earnings_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('LINEABOVE', (4,0), (4,1), 1, colors.black),
        ]))
        elements.append(Paragraph("Earnings", styles['Heading2']))
        elements.append(earnings_table)
        elements.append(Spacer(1, 0.3*inch))

        # Deductions Table
        deductions_data = [
            ["PF Deduction", f"₹{data['PF Deduction']}"],
            ["ESI Deduction", f"₹{data['ESI Deduction']}"],
            ["PT Deduction", f"₹{data['PT Deduction']}"],
            ["", ""],
            ["Total Deductions", f"₹{total_deductions}"]
        ]
        deductions_table = Table(deductions_data, colWidths=[3.5*inch, 1.5*inch])
        deductions_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('LINEABOVE', (4,0), (4,1), 1, colors.black),
        ]))
        elements.append(Paragraph("Deductions", styles['Heading2']))
        elements.append(deductions_table)
        elements.append(Spacer(1, 0.3*inch))

        # Net Pay Section
        net_pay_data = [
            ["Net Payable Amount", f"₹{net_pay}"]
        ]
        net_pay_table = Table(net_pay_data, colWidths=[3.5*inch, 1.5*inch])
        net_pay_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('BACKGROUND', (0,0), (-1,-1), colors.lightgrey),
        ]))
        elements.append(net_pay_table)

        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=2
        )
        elements.append(Paragraph("This is a computer generated document and does not require signature", footer_style))

        doc.build(elements)
        messagebox.showinfo("Success", f"PaySlip saved as:\n{filename}")
    except Exception as e:
        messagebox.showerror("PDF Error", f"Failed to generate PDF: {str(e)}")

def open_payroll_window():
    app.destroy()
    payroll_window = ctk.CTk()
    payroll_window.title("Employee Payroll Details")
    payroll_window.geometry("960x740")

    scrollable_frame = ctk.CTkScrollableFrame(payroll_window, fg_color="transparent")
    scrollable_frame.pack(fill="both", expand=True, padx=20, pady=20)

    try:
        book = load_workbook('data.xlsx')
        sheet = book.active

        fields = [
            ("Employee Name", "A"), ("Employee ID", "B"), ("Department", "C"),
            ("Designation", "D"), ("Date of Joining", "E"), ("Date of Birth", "F"),
            ("UAN", "G"), ("PF No", "H"), ("ESI No", "I"), ("Basic Salary", "J"),
            ("Conveyance", "K"), ("Special Allowance", "L"), ("PF Deduction", "M"),
            ("ESI Deduction", "N"), ("PT Deduction", "O"),
        ]

        for row in range(2, sheet.max_row + 1):
            data = {}
            for field, col in fields:
                cell = f"{col}{row}"
                data[field] = sheet[cell].value

            total_earnings = data["Basic Salary"] + data["Conveyance"] + data["Special Allowance"]
            total_deductions = data["PF Deduction"] + data["ESI Deduction"] + data["PT Deduction"]
            net_pay = total_earnings - total_deductions

            # Employee Frame
            employee_frame = ctk.CTkFrame(scrollable_frame, fg_color="#2E2E2E")
            employee_frame.pack(fill="x", padx=10, pady=10)

            # Details Column
            details_frame = ctk.CTkFrame(employee_frame, fg_color="transparent")
            details_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            ctk.CTkLabel(details_frame, text="Details", font=("Helvetica", 16, "bold")).pack(pady=10)
            
            details = [("Employee Name", data["Employee Name"]), ("Employee ID", data["Employee ID"]),
                      ("Department", data["Department"]), ("Designation", data["Designation"]),
                      ("Date of Joining", data["Date of Joining"]), ("Date of Birth", data["Date of Birth"]),
                      ("UAN", data["UAN"]), ("PF No", data["PF No"]), ("ESI No", data["ESI No"])]
            
            for field, value in details:
                row_frame = ctk.CTkFrame(details_frame, fg_color="transparent")
                row_frame.pack(fill="x", padx=10, pady=5)
                ctk.CTkLabel(row_frame, text=field, font=("Helvetica", 14)).pack(side="left", padx=10)
                ctk.CTkLabel(row_frame, text=value, font=("Helvetica", 14)).pack(side="right", padx=10)

            # Earnings Column
            earnings_frame = ctk.CTkFrame(employee_frame, fg_color="transparent")
            earnings_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            ctk.CTkLabel(earnings_frame, text="Earnings", font=("Helvetica", 16, "bold")).pack(pady=10)
            
            earnings = [("Basic Salary", f"₹{data['Basic Salary']}"), 
                       ("Conveyance", f"₹{data['Conveyance']}"), 
                       ("Special Allowance", f"₹{data['Special Allowance']}")]
            
            for field, value in earnings:
                row_frame = ctk.CTkFrame(earnings_frame, fg_color="transparent")
                row_frame.pack(fill="x", padx=10, pady=5)
                ctk.CTkLabel(row_frame, text=field, font=("Helvetica", 14)).pack(side="left", padx=10)
                ctk.CTkLabel(row_frame, text=value, font=("Helvetica", 14)).pack(side="right", padx=10)

            # Deductions Column
            deductions_frame = ctk.CTkFrame(employee_frame, fg_color="transparent")
            deductions_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            ctk.CTkLabel(deductions_frame, text="Deductions", font=("Helvetica", 16, "bold")).pack(pady=10)
            
            deductions = [("PF Deduction", f"₹{data['PF Deduction']}"), 
                         ("ESI Deduction", f"₹{data['ESI Deduction']}"), 
                         ("PT Deduction", f"₹{data['PT Deduction']}")]
            
            for field, value in deductions:
                row_frame = ctk.CTkFrame(deductions_frame, fg_color="transparent")
                row_frame.pack(fill="x", padx=10, pady=5)
                ctk.CTkLabel(row_frame, text=field, font=("Helvetica", 14)).pack(side="left", padx=10)
                ctk.CTkLabel(row_frame, text=value, font=("Helvetica", 14)).pack(side="right", padx=10)

            # Summary Section
            summary_frame = ctk.CTkFrame(employee_frame, fg_color="transparent")
            summary_frame.pack(fill="x", padx=20, pady=10)
            ctk.CTkLabel(summary_frame, text="Summary", font=("Helvetica", 18, "bold")).pack(pady=10)
            
            summary_data = [("Total Earnings", f"₹{total_earnings}"), 
                           ("Total Deductions", f"₹{total_deductions}"), 
                           ("Net Pay", f"₹{net_pay}")]
            
            for field, value in summary_data:
                row_frame = ctk.CTkFrame(summary_frame, fg_color="transparent")
                row_frame.pack(fill="x", padx=20, pady=5)
                ctk.CTkLabel(row_frame, text=field, font=("Helvetica", 16, "bold")).pack(side="left", padx=10)
                ctk.CTkLabel(row_frame, text=value, font=("Helvetica", 16, "bold")).pack(side="right", padx=10)

            # PDF Button
            print_button = ctk.CTkButton(
                employee_frame,
                text="Save as PDF",
                command=lambda d=data.copy(), te=total_earnings, td=total_deductions, np=net_pay: generate_pdf(d, te, td, np),
                font=("Helvetica", 14),
                fg_color="#27AE60",
                hover_color="#1E8449",
                text_color="white",
                corner_radius=8
            )
            print_button.pack(side="bottom", pady=15, padx=20, anchor="e")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load data: {str(e)}")

    payroll_window.mainloop()

# Main Application
app = ctk.CTk()
app.title("Employee Payroll Calculator")
app.geometry("400x200")

title_label = ctk.CTkLabel(app, text="Employee Payroll System", font=("Helvetica", 24, "bold"))
title_label.pack(pady=20)

calculate_button = ctk.CTkButton(
    app,
    text="Generate Payroll",
    command=open_payroll_window,
    font=("Helvetica", 16),
    fg_color="#2E86C1",
    hover_color="#1B4F72"
)
calculate_button.pack(pady=10)

app.mainloop()